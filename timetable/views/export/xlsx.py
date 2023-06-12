import itertools

from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from users.models.university import Auditorium
from timetable.models.timetable import TimetableItem
from datetime import datetime, timedelta


'''
header = [
    '111a', '122a', '123a', '222a', 
]
data = [
    ['para1', 'para1', 'para1', para1',],
    ['para2', 'para2', '',      para2',],
    ['para3', '',       'para3', para3',],
    ['',      'para4', 'para4', para4',],
]
'''

class ExportXlsxView(View):
    def get(self, request, monday, sunday):
        mon = datetime.strptime(monday, '%d_%m_%y').date()
        sun = datetime.strptime(sunday, '%d_%m_%y').date()

        classrooms = Auditorium.objects.values_list('name', flat=True).distinct().order_by('name')

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Schedule"

        headers = ["День недели"]
        headers.extend(classrooms)
        worksheet.append(headers)

        font = Font(name='Times New Roman', size=16)
        for cell in worksheet[1]:
            cell.font = font
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
        current_day = mon
        while current_day <= sun:
            schedule_rows = []

            for classroom_index, classroom in enumerate(classrooms):
                schedule_items = TimetableItem.objects.filter(date=current_day, auditorium__name=classroom).order_by(
                    'start_time', 'end_time')

                if schedule_items:
                    for item in schedule_items:
                        row_data = [f'day: \n{current_day.strftime("%d-%m-%Y")}']
                        row_data.extend(
                            ["" for _ in range(len(classrooms))])
                        row_data[classroom_index + 1] = f'{item.name}\n{item.start_time} - {item.end_time}'
                        schedule_rows.append(row_data)
                else:
                    row_data = [f'day: \n{current_day.strftime("%d-%m-%Y")}']
                    row_data.extend(["" for _ in range(len(classrooms))])
                    schedule_rows.append(row_data)
            for row_data in schedule_rows:
                worksheet.append(row_data)
            start_row = worksheet.max_row - len(schedule_rows) + 1
            end_row = worksheet.max_row
            worksheet.merge_cells(f'A{start_row}:A{end_row}')

            current_day += timedelta(days=1)

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.font = font
                cell.border = thin_border

        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=schedule.xlsx"
        workbook.save(response)
        return response
