"""
Module for creating a schedule in the classrooms of the ShSPU Technopark.
"""
import os
import openpyxl.styles
import datetime
import copy
from timetable.models.timetable import TimetableItem, Type_TimetableItem
from users.models.university import Auditorium, University_Unit


def json_to_excel(university_id, monday: str, sunday: str):

    json_list = Auditorium.objects.filter(
        university_unit__show_in_timetable=True,
        university_unit_id=university_id
    ).all()
    start_week = datetime.datetime.strptime(monday, '%d_%m_%y').date()
    end_week = datetime.datetime.strptime(sunday, '%d_%m_%y').date()
    json_couples = TimetableItem.objects.filter(
            auditorium__university_unit_id=university_id,
            date__range=[start_week, end_week],
            status='APPROVED',
        ).all().order_by('date', 'start_time')
    weekdays = (
        "ПН",
        "ВТ",
        "СР",
        "ЧТ",
        "ПТ",
        "СБ",
    )
    times = (
        "8:00-9:30",
        "9:40-11:10",
        "11:20-12:50",
        "13:20-14:50",
        "15:00-16:30",
        "16:40-18:10",
    )
    couples_indexes = {}
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    teacher_name = ''
    let_num = 2
    row_num = 3
    times_index = 0
    path = os.getcwd()
    sym_count = 0
    skip = False
    auditoriums_num = len(json_list)
    schedule = openpyxl.Workbook()
    schedule.remove(schedule.active)
    sheet = schedule.create_sheet("Расписание")
    sheet.freeze_panes = "C3"
    sheet.row_dimensions[1].height = 25
    sheet.row_dimensions[2].height = 70
    for auditorium in json_list:
        sheet.column_dimensions[alphabet[let_num]].width = 60
        couples_indexes[auditorium.name] = []
        sheet[f"{alphabet[let_num]}1"] = auditorium.name
        sheet[f"{alphabet[let_num]}1"].font = openpyxl.styles.Font(name="Times New Roman", size=22, bold=True)
        sheet[f"{alphabet[let_num]}1"].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        sheet[f"{alphabet[let_num]}1"].border = openpyxl.styles.Border(left=openpyxl.styles.Side(style="thin"))
        sheet[f"{alphabet[let_num]}2"] = auditorium.info
        sheet[f"{alphabet[let_num]}2"].font = openpyxl.styles.Font(name="Times New Roman", size=12)
        sheet[f"{alphabet[let_num]}2"].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center",
                                                                             wrap_text=True)
        sheet[f"{alphabet[let_num]}2"].border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style="thick"),
                                                                       left=openpyxl.styles.Side(style="thin"))
        if let_num == 2:
            sheet["B1"].border = openpyxl.styles.Border(right=openpyxl.styles.Side(style="thick"))
        elif len(json_list) == let_num - 1:
            sheet.column_dimensions[alphabet[let_num + 1]].width = 50
            sheet[f"{alphabet[let_num + 1]}1"] = "Мероприятия"
            sheet[f"{alphabet[let_num + 1]}1"].font = openpyxl.styles.Font(name="Times New Roman", size=22, bold=True)
            sheet[f"{alphabet[let_num + 1]}1"].alignment = openpyxl.styles.Alignment(horizontal="center",
                                                                                     vertical="center")
            sheet[f"{alphabet[let_num + 1]}1"].border = openpyxl.styles.Border(
                right=openpyxl.styles.Side(style="thick"),
                left=openpyxl.styles.Side(style="thin"))
            sheet[f"{alphabet[let_num + 1]}2"].border = openpyxl.styles.Border(
                bottom=openpyxl.styles.Side(style="thick"),
                right=openpyxl.styles.Side(style="thick"),
                left=openpyxl.styles.Side(style="thin"))
        let_num += 1
    aud_index = let_num
    sheet.column_dimensions['A'].width = 15
    sheet[f"A{row_num}"].border = openpyxl.styles.Border(top=openpyxl.styles.Side(style="thick"))
    date = start_week
    for day in weekdays:
        sheet[f"A{row_num}"] = day
        sheet[f"A{row_num + 2}"] = f"{date.day}.{date.month}.{date.year}"
        sheet[f"A{row_num}"].font = openpyxl.styles.Font(name="Times New Roman", size=14)
        sheet[f"A{row_num + 2}"].font = openpyxl.styles.Font(name="Times New Roman", size=14)
        row_num += 13
        sheet[f"A{row_num - 1}"].border = openpyxl.styles.Border(top=openpyxl.styles.Side(style="thick"),
                                                                 bottom=openpyxl.styles.Side(style="thick"))
        sheet[f"A{row_num - 1}"].fill = openpyxl.styles.PatternFill(start_color="C0C0C0", fill_type="solid")
        date += datetime.timedelta(1)
    sheet["A1"].border = openpyxl.styles.Border(right=openpyxl.styles.Side(style="thin"))
    sheet["B2"].border = openpyxl.styles.Border(left=openpyxl.styles.Side(style="thin"),
                                                right=openpyxl.styles.Side(style="thick"),
                                                bottom=openpyxl.styles.Side(style="thick"))
    sheet.column_dimensions['B'].width = 15
    for row_num in range(3, 81):
        if times_index == 6 and not skip:
            times_index = 0
            sheet[f"B{row_num}"].border = openpyxl.styles.Border(top=openpyxl.styles.Side(style="thick"),
                                                                 bottom=openpyxl.styles.Side(style="thick"))
            sheet[f"B{row_num}"].fill = openpyxl.styles.PatternFill(start_color="C0C0C0", fill_type="solid")
            continue
        elif skip:
            sheet[f"B{row_num}"].border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style="thin"),
                                                                 left=openpyxl.styles.Side(style="thin"),
                                                                 right=openpyxl.styles.Side(style="thick"))
            skip = False
            continue
        sheet[f"B{row_num}"] = times[times_index]
        sheet[f"B{row_num}"].font = openpyxl.styles.Font(name="Times New Roman", size=14)
        sheet[f"B{row_num}"].border = openpyxl.styles.Border(left=openpyxl.styles.Side(style="thin"),
                                                             right=openpyxl.styles.Side(style="thick"))
        skip = True
        times_index += 1
    for info in json_couples:
        if info.auditorium.all()[0].name in couples_indexes:
            couples_indexes[info.auditorium.all()[0].name].append(list(json_couples).index(info))
    couples_indexes_copy = copy.deepcopy(couples_indexes)
    for key, value in couples_indexes.items():
        for index_value in value:
            info = json_couples[index_value].name
            if info.rfind('.') != -1 and info.rfind('.') != len(info) - 1:
                teacher_name = info[:info.rfind('.') + 1]
                info = (f"{info[info.rfind('.') + 2:].replace(f' {key}', '')}"
                        f"{json_couples[index_value].info.split('группы')[1]}")
            elif info.rfind('.') == -1:
                if json_couples[index_value].info.count('.') == 2:
                    teacher_name = json_couples[index_value].info
                else:
                    teacher_name = ''
            else:
                for index_info in range(len(info) - 1, -1, -1):
                    if info[index_info] == ' ':
                        sym_count += 1
                    if sym_count == 2:
                        sym_count = 0
                        teacher_name = info[index_info + 1:len(info)]
                        info = info[:index_info - 2].replace(f" {key}", '')
                        break
                if json_couples[index_value]["type"]["name"] != "Мероприятие":
                    info += json_couples[index_value]["info"].split("группы")[1]
            for let_num in range(len(couples_indexes)):
                if sheet[f"{alphabet[let_num + 2]}1"].value == key:
                    times_index = 0
                    day_index = 5
                    skip = False
                    for row_num in range(3, 79):
                        if times_index == 6 and not skip:
                            times_index = 0
                            day_index += 13
                            continue
                        elif skip:
                            skip = False
                            continue
                        skip = True
                        times_index += 1
                        excel_date = sheet[f"A{day_index}"].value.split('.')
                        data_date = str(json_couples[index_value].date).split('-')
                        excel_start_time = datetime.time(int(sheet[f"B{row_num}"].value.split('-')[0].split(':')[0]),
                                                         int(sheet[f"B{row_num}"].value.split('-')[0].split(':')[1]))
                        data_start_time = datetime.time(
                            int(str(json_couples[index_value].start_time).split(':')[0]),
                            int(str(json_couples[index_value].start_time).split(':')[1])
                        )
                        excel_end_time = datetime.time(int(sheet[f"B{row_num}"].value.split('-')[1].split(':')[0]),
                                                       int(sheet[f"B{row_num}"].value.split('-')[1].split(':')[1]))
                        data_end_time = datetime.time(
                            int(str(json_couples[index_value].end_time).split(':')[0]),
                            int(str(json_couples[index_value].end_time).split(':')[1])
                        )
                        if (datetime.date(int(excel_date[2]), int(excel_date[1]), int(excel_date[0]))
                                == datetime.date(int(data_date[0]), int(data_date[1]), int(data_date[2]))
                                and excel_start_time == data_start_time and excel_end_time == data_end_time
                                and sheet[f"{alphabet[let_num + 2]}{row_num}"].value is None
                                and sheet[f"{alphabet[let_num + 2]}{row_num + 1}"].value is None):
                            sheet[f"{alphabet[let_num + 2]}{row_num}"] = teacher_name
                            sheet[f"{alphabet[let_num + 2]}{row_num + 1}"] = info
                            couples_indexes_copy[key].remove(index_value)
                            sheet[f"{alphabet[let_num + 2]}{row_num}"].font = openpyxl.styles.Font(
                                name="Times New Roman", size=12)
                            sheet[f"{alphabet[let_num + 2]}{row_num + 1}"].font = openpyxl.styles.Font(
                                name="Times New Roman", size=12)
                            sheet[f"{alphabet[let_num + 2]}{row_num}"].alignment = openpyxl.styles.Alignment(
                                horizontal="center", vertical="center")
                            sheet[f"{alphabet[let_num + 2]}{row_num + 1}"].alignment = openpyxl.styles.Alignment(
                                horizontal="center", vertical="center")
                    break
    couples_indexes = copy.deepcopy(couples_indexes_copy)
    for key, value in couples_indexes.items():
        for index_value in value:
            info = json_couples[index_value].name
            for let_num in range(len(couples_indexes)):
                if sheet[f"{alphabet[let_num + 2]}1"].value == key:
                    times_index = 0
                    day_index = 5
                    skip = False
                    for row_num in range(3, 79):
                        if times_index > 5 and not skip:
                            times_index = 0
                            day_index += 13
                            continue
                        elif skip:
                            skip = False
                            continue
                        skip = True
                        times_index += 1
                        excel_date = sheet[f"A{day_index}"].value.split('.')
                        data_date = str(json_couples[index_value].date).split('-')
                        data_end_time = datetime.time(
                            int(str(json_couples[index_value].end_time).split(':')[0]),
                            int(str(json_couples[index_value].end_time).split(':')[1])
                        )
                        excel_end_time = datetime.time(int(sheet[f"B{row_num}"].value.split('-')[1].split(':')[0]),
                                                       int(sheet[f"B{row_num}"].value.split('-')[1].split(':')[1]))
                        if (datetime.date(int(excel_date[2]), int(excel_date[1]), int(excel_date[0]))
                                == datetime.date(int(data_date[0]), int(data_date[1]), int(data_date[2]))
                                and sheet[f"{alphabet[let_num + 2]}{row_num}"].value is None
                                and sheet[f"{alphabet[let_num + 2]}{row_num + 1}"].value is None
                                and excel_end_time >= data_end_time and index_value in couples_indexes_copy[key]):
                            sheet[f"{alphabet[let_num + 2]}{row_num}"] = \
                                (f"{key} {str(json_couples[index_value].start_time)[:-3]}-"
                                 f"{str(json_couples[index_value].end_time)[:-3]}")
                            sheet[f"{alphabet[let_num + 2]}{row_num + 1}"] = info
                            couples_indexes_copy[key].remove(index_value)
                            sheet[f"{alphabet[let_num + 2]}{row_num}"].font = openpyxl.styles.Font(
                                name="Times New Roman", size=12, bold=True)
                            sheet[f"{alphabet[let_num + 2]}{row_num + 1}"].font = openpyxl.styles.Font(
                                name="Times New Roman", size=12, bold=True)
                            sheet[f"{alphabet[let_num + 2]}{row_num}"].alignment = openpyxl.styles.Alignment(
                                horizontal="center", vertical="center")
                            sheet[f"{alphabet[let_num + 2]}{row_num + 1}"].alignment = openpyxl.styles.Alignment(
                                horizontal="center", vertical="center")
                    break
    couples_indexes = copy.deepcopy(couples_indexes_copy)
    for key, value in couples_indexes.items():
        for index_value in value:
            info = json_couples[index_value]["name"]
            times_index = 0
            day_index = 5
            skip = False
            for row_num in range(3, 79):
                if times_index >= 6 and not skip:
                    times_index = 0
                    day_index += 13
                    continue
                elif skip:
                    skip = False
                    continue
                skip = True
                times_index += 1
                excel_date = sheet[f"A{day_index}"].value.split('.')
                data_date = json_couples[index_value]["date"].split('-')
                if (datetime.date(int(excel_date[2]), int(excel_date[1]), int(excel_date[0]))
                        == datetime.date(int(data_date[2]), int(data_date[1]), int(data_date[0]))
                        and sheet[f"{alphabet[aud_index]}{row_num}"].value is None
                        and sheet[f"{alphabet[aud_index]}{row_num + 1}"].value is None
                        and index_value in couples_indexes_copy[key]):
                    sheet[f"{alphabet[aud_index]}{row_num}"] = \
                        (f"{key} {json_couples[index_value]['start_time'][:-3]}-"
                         f"{json_couples[index_value]['end_time'][:-3]}")
                    sheet[f"{alphabet[aud_index]}{row_num + 1}"] = info
                    sheet[f"{alphabet[aud_index]}{row_num}"].font = openpyxl.styles.Font(name="Times New Roman",
                                                                                         size=12)
                    sheet[f"{alphabet[aud_index]}{row_num}"].alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center")
                    sheet[f"{alphabet[aud_index]}{row_num + 1}"].font = openpyxl.styles.Font(
                        name="Times New Roman", size=12)
                    sheet[f"{alphabet[aud_index]}{row_num + 1}"].alignment = openpyxl.styles.Alignment(
                        horizontal="center", vertical="center")
                    couples_indexes_copy[key].remove(index_value)
                    break
    for let_num in range(2, auditoriums_num + 3):
        times_index = 0
        skip = False
        for row_num in range(3, 81):
            if let_num == auditoriums_num + 2:
                sheet[f"{alphabet[let_num + 1]}{row_num}"].border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style="thick"))
            if times_index == 6 and not skip:
                times_index = 0
                sheet[f"{alphabet[let_num]}{row_num}"].border = openpyxl.styles.Border(
                    top=openpyxl.styles.Side(style="thick"),
                    bottom=openpyxl.styles.Side(style="thick"))
                sheet[f"{alphabet[let_num]}{row_num}"].fill = openpyxl.styles.PatternFill(start_color="C0C0C0",
                                                                                          fill_type="solid")
                continue
            elif skip:
                sheet[f"{alphabet[let_num]}{row_num}"].border = openpyxl.styles.Border(
                    bottom=openpyxl.styles.Side(style="thin"),
                    right=openpyxl.styles.Side(style="thin"))
                skip = False
                continue
            sheet[f"{alphabet[let_num]}{row_num}"].border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style="thin"),
                right=openpyxl.styles.Side(style="thin"))
            skip = True
            times_index += 1
    file_name = \
        f"Technopark_schedule_{sheet['A5'].value.replace('.', '_')}_"\
        f"{sheet['A70'].value.replace('.', '_')}.xlsx"
    schedule.save(path + '/' + file_name)

    return file_name
